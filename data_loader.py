import os
import pytz
import pandas as pd
import cPickle
from datetime import datetime, timedelta
from openpyxl.reader.excel import load_workbook
from sklearn import datasets as ds
from Preprocessing import *

def time_processing(t, tz1, tz2, fmt):
    '''
    transfer time t at timezone tz1 to timezone tz2
    :param t: time t to be transferred
    :param tz1: timezone info of time t
    :param tz2: timezone after transfer
    :param fmt:
    :return:
    '''
    YY, MM, DD, hh, mm, ss = int(t[:4]), int(t[5:7]), int(t[8:10]), int(t[11:13]), int(t[14:16]), int(t[17:19])
    utc_dt = datetime(YY, MM, DD, hh, mm, ss, tzinfo=tz1)
    est = utc_dt.astimezone(tz2)
    t = est.strftime(fmt)
    # preprocessing time to US/Eastern timezone, only show date
    # if the message is posted before market open, mark it previous date
    # (next trading day to be considered)
    if (t.endswith('EDT-0400') and (t[11:19] >= '09:25:00')) or (
                t.endswith('EDT-0500') and (t[11:19] >= '10:25:00')):
        t = t[:10]
    else:
        t = (est - timedelta(days=1)).strftime(fmt)[:10]
    return t

def day_diff(t1, t2):
    '''
    day diff between t1 and t2 (t1<t2)
    :param t1: format 'YYYY-MM-DD'
    :param t2: format 'YYYY-MM-DD'
    :return: #day_diff
    '''
    t1 = datetime(int(t1[:4]), int(t1[5:7]), int(t1[8:10]))
    t2 = datetime(int(t2[:4]), int(t2[5:7]), int(t2[8:10]))
    diff = t2 - t1
    return diff.days

def trend(target_price, basic_price, threshold = 0.01):
    '''
    :param target_price: price after time interval
    :param basic_price: price when predicting
    :param threshold: threshold of price changes
    :return: label of trend. 0: stay; 1: down; 2: up
    '''
    percent = (float(target_price) -  float(basic_price)) / float(basic_price)
    if percent >= threshold:
        return 1
    elif percent <= -threshold:
        return -1
    else:
        return 0

def get_price_label(price_data, t, time_interval = 1, threshold = 0.03):
    fmt = '%Y-%m-%d'
    idx = price_data.index
    target_time = (datetime.strptime(t, fmt) + timedelta(days=time_interval)).strftime(fmt)
    # if market is not open at t, then t-1
    while t not in idx:
        t = (datetime.strptime(t, fmt) - timedelta(days=1)).strftime(fmt)
    # if market is not open at target_time, then target_time+1
    while target_time not in idx:
        target_time = (datetime.strptime(target_time, fmt) + timedelta(days=1)).strftime(fmt)

    # set label
    label = trend(price_data[target_time], price_data[t], threshold)
    return label


def file_processing(symbol, rate = '6:2:2',train_data = dict(), hold_data = dict(), test_data = dict(), time_interval = 1, msg_dir = 'stocktwits_samples/', price_dir = 'stock_prices/'):
    # read msg file
    msg_file = msg_dir + symbol + '.xlsx'
    wb = load_workbook(filename=msg_file)
    ws = wb.get_active_sheet()
    max_row, max_col = ws.max_row, ws.max_column
    print 'For', symbol, ', max_row = ', max_row

    col_dict = {1: 'A', 2: 'B', 3: 'C', 4: 'D', 5: 'E', 6: 'F', 7: 'G', 8: 'H', 9: 'I', 10: 'J', 11: 'K', 12: 'L',
                13: 'M', 14: 'N', 15: 'O', 16: 'P', 17: 'Q', 18: 'R', 19: 'S', 20: 'T', 21: 'U', 22: 'V', 23: 'W',
                24: 'X', 25: 'Y', 26: 'Z'}
    # match hearders and corresponding column
    headers = dict()
    for c in range(1, max_col+1):
        s = ws[col_dict[c]+'1'].value.encode('utf-8')
        headers[s] = col_dict[c]

    # transfer all punctuations to whitespace
    translator = string.maketrans(string.punctuation, " " * len(string.punctuation))

    # transfer timezone from utc to US/Eastern timezone
    utc = pytz.utc
    eastern = pytz.timezone('US/Eastern')
    fmt = '%Y-%m-%d %H:%M:%S %Z%z'

    # read price file
    price_file = price_dir + symbol + '.csv'
    if not os.path.exists(price_file):
        print 'price file doesn\'t exist. Please input correct directory or price file name.'
        return
    dateparse = lambda dates: pd.datetime.strptime(dates, '%Y-%m-%d')
    price_data = pd.read_csv(price_file, parse_dates=[0], index_col='Date', date_parser=dateparse, usecols=[0,1])   # the usecols indicates which column to be used
    price_data = price_data['Open'] # use open price, other price: High, Low, Close, Adj Close, Volume


    # save data
    row = [int(k) for k in rate.split(':')]
    row = [int(i * max_row/sum(row)) for i in row[:-1]]
    row[1] += row[0]
    row = [max_row - i for i in row]    # reverse as time sequence

    for i in range(max_row, 1, -1):
        # read summary and save it into the msg_list
        one_msg = dict()
        msg = ws[headers['Message']+str(i)].value  # message
        if msg is not None:
            #context = ws[headers['Context']+str(i)].value
            #if context is not None:
            #    msg += context
            one_msg["content"] = msg.encode('utf-8').replace("$"+symbol, " ")
        else:
            continue

        # read time and transfer it to US/Eastern Timezone and save it into the time_info
        post_date = time_processing(ws[headers['Created_time']+str(i)].value.encode('utf-8'), utc, eastern, fmt)
        one_msg["time"] = post_date # only save the date
        label = get_price_label(price_data, post_date, time_interval, threshold=0.03)
        one_msg["label"] = label
        fluctuation = get_price_label(price_data, post_date, time_interval, threshold=0.01)
        one_msg["fluctuation"] = abs(fluctuation)

        # read sentiment
        senti = ws[headers['Sentiment']+str(i)].value
        if senti is not None:
            if senti.encode('utf-8') == "Bullish":
                one_msg["sentiment"] = 1
            elif senti.encode('utf-8') == "Bearish":
                one_msg["sentiment"] = -1
        else:
            one_msg["sentiment"] = 0

        # msg_like_count
        one_msg["msg_like_count"] = int(ws[headers['Msg_Like_Count']+str(i)].value)

        # read user information and save it one_user
        one_user = dict()
        userid = int(ws[headers['UserID']+str(i)].value)

        if not ws[headers['Official']+str(i)].value:
            one_user["official"] = 0
        else:
            one_user["official"] = 1
        classification = ws[headers['Classification']+str(i)].value
        if classification is not None and classification.encode('utf-8').lower().find('suggested') >= 0:
            one_user["suggested"] = 1
        else:
            one_user["suggested"] = 0

        one_user["user_like_count"] = int(ws[headers['User_Like_Count']+str(i)].value)
        one_user["ideas"] = int(ws[headers['Ideas']+str(i)].value)
        one_user["followers"] = int(ws[headers['Followers']+str(i)].value)
        one_user["following"] = int(ws[headers['Following']+str(i)].value)

        one_msg["join_days"] = day_diff(ws[headers['Join_Date']+str(i)].value.encode('utf-8'), post_date)

        if post_date < '2015': # i >= row[0]:
            if train_data.has_key(userid):
                train_data[userid]["messages"].append(one_msg)
            else:
                train_data[userid] = dict()
                train_data[userid]["messages"] = [one_msg]
                train_data[userid]["user_info"] = one_user
        elif post_date < '2017': # i >= row[1]:
            if hold_data.has_key(userid):
                hold_data[userid]["messages"].append(one_msg)
            else:
                hold_data[userid] = dict()
                hold_data[userid]["messages"] = [one_msg]
                hold_data[userid]["user_info"] = one_user
        else:
            if test_data.has_key(userid):
                test_data[userid]["messages"].append(one_msg)
            else:
                test_data[userid] = dict()
                test_data[userid]["messages"] = [one_msg]
                test_data[userid]["user_info"] = one_user

    return train_data, hold_data, test_data

def msg_process(one_msg):
    msg = one_msg["content"]
    existence = [0] * 5 # url, number, hashtag, question_mark, great_fluctuation(>=1%)

    # url
    while True:
        url_idx = msg.find("http")
        if url_idx >= 0:
            existence[0] = 1
            nxt_space = msg.find(' ', url_idx)
            if nxt_space >= 0:
                # remove url
                msg = msg[:url_idx] + msg[nxt_space+1:]
            else:
                msg = msg[:url_idx]
        else:
            break

    # numbers
    idxs = re.findall(r'\d+', msg)
    if len(idxs):
        existence[1] = 1
        for n in idxs:
            msg.replace(n,' ')

    # hashtag
    while True:
        idxs = re.findall('#', msg)
        if len(idxs) >= 2:
            existence[2] = 1
            idx1 = msg.find("#")
            idx2 = msg.find("#", idx1)
            msg = msg[:idx1] + msg[idx2 + 1:]
        else:
            break

    # question mark
    qm_idx = msg.find("?")
    if qm_idx >= 0:
        existence[3] = 1
    msg.replace("?", " ")

    # whether fluctuate no less than 1% in the previous trading day
    existence[4] = one_msg["fluctuation"]

    return msg, existence

def dataset_process(dataset, type):
    user_list = dict()  # {userid: #msg}
    num_words = len(vocab)
    before_count = 0 # the #data before current user
    for userid, info in dataset.items():
        msgs = info["messages"] # list of messages
        user = info["user_info"]
        tmp = list(map(lambda x: msg_process(x), msgs))
        pro_msg = list(map(lambda x:x[0], tmp))
        existences = list(map(lambda x:x[1], tmp))
        labels = map(lambda x: x["label"], msgs)
        try:
            words_idx = Article2Index(pro_msg, vocab, padding=False)
        except:
            print 'Error msg = ',pro_msg

        s = ""
        count = 0
        for i in range(len(msgs)):
            if labels[i] == 0:
                continue
            # label
            s += str(labels[i]) + ' '
            # words' index
            idx = sorted(set(words_idx[i]))
            for e in idx:
                s += str(e) +':1 '
            # existence of some elements
            exist = existences[i]
            for j,e in enumerate(exist):
                s += str(j+num_words) + ':' + str(e) + ' '
            if s != "":
                s += '\n'
                count += 1
        #user_list.append(tuple((userid, count)))
        user_list[userid] = (count, before_count)
        before_count += count

        # store idxes into files
        with open('mid-data/'+type +'.txt','a') as f:
            f.write(s)

    with open('mid-data/UserInfo.pkl', 'a') as f:
        cPickle.dump(user_list, f)

def data_loader(symbols, rate = '3:2:2', time_interval = 1, msg_dir = "stocktwits_samples/", price_dir = "stock_prices/", min_freq = 5):
    '''
    :param symbol: symbol of the stock
    :param model: fasttext model
    :param directory: directory of file saving summary and price
    :param time_interval: prediction time interval
    :return: train_dataset, test_dataset consisting of tuples of (inputs, targets)
    '''
    stopwords_list = []

    # get vocab
    vocab_file = 'vocab_' + str(min_freq) + '.pkl'
    global vocab
    if os.path.exists(vocab_file):
        with open(vocab_file,'r') as f:
            vocab = cPickle.load(f)
    else:
        train_file = 'alltext.txt'
        if os.path.exists(train_file):
            with open(train_file, 'r') as f:
               train_texts = f.read()
            vocab = getVocab(train_texts, min_freq, stopwords_list)

        else:
            print 'cannot find the training text file, thus failing to generate vocabulary. Please try again.'
            return

    # get datasets for given symbol
    kinds = ['train', 'hold', 'test']
    data_files = ['mid-data/' + type + '.txt' for type in kinds]
    user_file = 'mid-data/UserInfo.pkl'
    # initialize dataset
    train_set = dict()
    hold_set = dict()
    test_set = dict()
    if not (os.path.exists(data_files[0]) and os.path.exists(data_files[1]) and os.path.exists(data_files[2]) and os.path.exists(user_file)):
        for symbol in symbols:
            # process and return dataset
            train_set, hold_set, test_set = file_processing(symbol=symbol, rate=rate, train_data=train_set, hold_data=hold_set, test_data=test_set,time_interval=time_interval, msg_dir=msg_dir, price_dir=price_dir)
        # save dataset and user information into seperate files
        all_datasets = (train_set, hold_set, test_set)
        print 'Before process, #users in each dataset: #train = %d, #hold = %d, test = %d' % (len(train_set), len(hold_set), len(test_set))
        map(lambda x, y: dataset_process(x, y), all_datasets, kinds)

    # load data from local files
    x_train, y_train, x_hold, y_hold, x_test, y_test = ds.load_svmlight_files(data_files)
    print 'After process: #train = %d, #hold = %d, test = %d' % (x_train.shape[0], x_hold.shape[0], x_test.shape[0])
    train_data = (x_train, y_train)
    hold_data = (x_hold, y_hold)
    test_data = (x_test, y_test)
    with open(user_file, 'r') as f:
        user_train = cPickle.load(f)
        user_hold = cPickle.load(f)
        user_test = cPickle.load(f)

    return (train_data, user_train), (hold_data, user_hold), (test_data, user_test)

